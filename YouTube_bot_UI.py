#!/usr/local/bin/python
# coding: utf-8
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle
import datetime
import sys
from pprint import pprint
import json
import re
import traceback
from pytube import YouTube
import os
import functools
from threading import Thread, Semaphore
import time

'''
Бот для взаимодействия с YouTube.
При запуске всплывает диалоговое окно, в которое нужно ввести название интересующего канала.
После нужно подождать некоторое время пока информация о канале, его плейлистах и канала появиться на таблицах интерфейса.
Можно скачать эти таблици в файл .exel
Также можно выделить интересующие видео (если хотите выделить все видео, то поставьте галочку под полем "скачать" во вкладке с информацией о профиле,
аналогично можно выделить все видео из плейлиста или отдельное видео) и затем скачать их, нажав на кнопку скачать видео.
Затем в папке с ботом будет создана папка с названием канала в которую и загрузятся все выделенные видео.
'''

def error_decorator(func):
    @functools.wraps(func)
    def wraper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except:
            traceback.print_exc(file=sys.stdout)
    return wraper

class first_dialog(QDialog):

    def __init__(self):
        super().__init__()

        self.setWindowTitle('YouTubeBot')
        self.setWindowIcon(QIcon('youtube_icon.jpg'))

        self.text = QLineEdit(self)
        self.text.setFixedHeight(65)
        self.text.setPlaceholderText('Davie504')
        self.text.setAlignment(Qt.AlignCenter)
        self.text.setFocusPolicy(Qt.ClickFocus)

        self.button = QPushButton('Далее', self)
        self.button.setFixedWidth(100)

        self.label = QLabel('Введите имя пользователя', self)
        self.label.setAlignment(Qt.AlignCenter)

        self.h_layout = QHBoxLayout()
        self.h_layout.addSpacing(0)
        self.h_layout.addWidget(self.button)

        self.layout = QVBoxLayout()
        self.layout.addWidget(self.label)
        self.layout.addSpacing(50)
        self.layout.addWidget(self.text)
        self.layout.addSpacing(25)
        self.layout.addLayout(self.h_layout)
        self.layout.addSpacing(40)
        self.setLayout(self.layout)

        self.setFixedSize(500, 270)
        self.setStyleSheet('''
        QDialog{
            background: #ffffff;
        }
        QLabel{
            font-size: 24px;
            color: #ce2c2a;
            font-weight: bold;
        }
        QLineEdit{
            font-size: 20px;
            color: #ce2c2a;
            selection-background-color: #ce2c2a;
            selection-color: #ffffff;
            border: 1px solid #ce2c2a;
            border-radius: 10px;
            margin-bottom: 30px;
        }
        QPushButton{
            font-size: 20px;
            background: #ffffff;
            color: #ce2c2a;
            border: 2px solid #ce2c2a;
            border-radius: 6px;
            padding: 10px 0px 10px 0px;
        }
        QPushButton:hover{
            background: #ce2c2a;
            color: #ffffff;
        }
        ''')

class my_tab_widget(QTabWidget):

    def __init__(self, parent):
        super().__init__(parent)

        self.chack_box_states = {}

        self.table_user = QTableWidget()
        self.table_playlist = QTableWidget()
        self.table_video = QTableWidget()

        self.table_user.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_playlist.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_video.setEditTriggers(QTableWidget.NoEditTriggers)

        self.addTab(self.table_user, 'Аккаунт')
        self.addTab(self.table_playlist, 'Плейлисты')
        self.addTab(self.table_video, 'Видео')

    def info_to_table(self, table, info, chanel_name):

        self.chanel_name = chanel_name
        special_index = float('inf')

        if table == self.table_user:
            columns = len(info) + 1
            rows = 1

            table.setColumnCount(columns)
            table.setRowCount(rows)

            for index, headers in enumerate(info):
                item = QTableWidgetItem()
                item.setText(headers)
                item.setTextAlignment(Qt.AlignCenter)
                table.setHorizontalHeaderItem(index, item)
                if headers == 'Описание':
                        special_index = index

            for index, headers in enumerate(info):
                data = str(info[headers])
                time = re.search(r'(\d+-\d+-\d+T\d+:\d+:\d+)', data)

                if time:
                    data = str(time.group(0).replace('T', ' '))

                elif data.isdigit() and len(data) > 3:
                    new_str = ''
                    for index_, number in enumerate(data[::-1]):
                        order = index_ + 1
                        if order % 3 == 0:
                            new_str += number + ','
                        else:
                            new_str += number
                    data = new_str.strip()[::-1]

                    if data[0] == ',':
                        data = data[1:]

                if index == special_index:
                    with open('a.txt', 'w', encoding='utf-8') as file:
                        pprint(data, stream=file)
                    with open('a.txt', 'r', encoding='utf-8') as file:
                        data = file.read()[1:-2].replace("'", '').replace(r'\n', '')

                item = QTableWidgetItem()
                item.setText(data)
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(0, index, item)

        elif table == self.table_video:
            try:
                rows = len(info)
                for item_number in info:
                    columns = len(info[item_number]) + 2

                table.setColumnCount(columns)
                table.setRowCount(rows)

                for index, headers in enumerate(info[item_number]):
                    item = QTableWidgetItem()
                    item.setText(headers)
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setHorizontalHeaderItem(index, item)
                    if headers == 'Описание':
                        special_index = index

                for item_number in info:
                    row = int(item_number) - 1
                    for index, header in enumerate(info[item_number]):
                        data = str(info[item_number][header])

                        time = re.search(r'(\d+-\d+-\d+T\d+:\d+:\d+)', data)

                        if time:
                            data = str(time.group(0).replace('T', ' '))

                        elif data.isdigit() and len(data) > 3:
                            new_str = ''
                            for index_, number in enumerate(data[::-1]):
                                order = index_ + 1
                                if order % 3 == 0:
                                    new_str += number + ','
                                else:
                                    new_str += number
                            data = new_str.strip()[::-1]
                            if data[0] == ',':
                                data = data[1:]

                        if index == special_index:
                            with open('a.txt', 'w', encoding='utf-8') as file:
                                pprint(data, stream=file)
                            with open('a.txt', 'r', encoding='utf-8') as file:
                                data = file.read()[1:-2].replace("'", '').replace(r'\n', '')

                        item = QTableWidgetItem()
                        item.setText(data)
                        item.setTextAlignment(Qt.AlignCenter)
                        table.setItem(row, index, item)
            except:
                traceback.print_exc(file=sys.stdout)

        else:
            try:
                rows = len(info)
                for item_number in info:
                    columns = len(info[item_number]) + 1

                table.setColumnCount(columns)
                table.setRowCount(rows)

                for index, headers in enumerate(info[item_number]):
                    item = QTableWidgetItem()
                    item.setText(headers)
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setHorizontalHeaderItem(index, item)
                    if headers == 'Описание':
                        special_index = index

                for item_number in info:
                    row = int(item_number) - 1
                    for index, header in enumerate(info[item_number]):
                        data = str(info[item_number][header])

                        time = re.search(r'(\d+-\d+-\d+T\d+:\d+:\d+)', data)

                        if time:
                            data = str(time.group(0).replace('T', ' '))

                        elif data.isdigit() and len(data) > 3:
                            new_str = ''
                            for index_, number in enumerate(data[::-1]):
                                order = index_ + 1
                                if order % 3 == 0:
                                    new_str += number + ','
                                else:
                                    new_str += number
                            data = new_str.strip()[::-1]
                            if data[0] == ',':
                                data = data[1:]

                        if index == special_index:
                            with open('a.txt', 'w', encoding='utf-8') as file:
                                pprint(data, stream=file)
                            with open('a.txt', 'r', encoding='utf-8') as file:
                                data = file.read()[1:-2].replace("'", '').replace(r'\n', '')

                        item = QTableWidgetItem()
                        item.setText(data)
                        item.setTextAlignment(Qt.AlignCenter)
                        table.setItem(row, index, item)
            except:
                traceback.print_exc(file=sys.stdout)

        item = QTableWidgetItem()
        item.setText('Скачать')
        item.setTextAlignment(Qt.AlignCenter)
        table.setHorizontalHeaderItem(columns - 1, item)

        for row in range(rows):
            check_box = QCheckBox(table)
            self.chack_box_states[check_box] = 0
            check_box.stateChanged.connect(lambda state: self.auto_check(table, state))
            check_box.setStyleSheet('QCheckBox{padding-left:20px}')
            table.setCellWidget(row, columns - 1, check_box)

        if table == self.table_video:
            item = QTableWidgetItem()
            item.setText('Прогресс')
            item.setTextAlignment(Qt.AlignCenter)
            table.setHorizontalHeaderItem(columns - 2, item)

            for row in range(rows):
                progress_bar = QProgressBar(table)
                table.setCellWidget(row, columns - 2, progress_bar)
                video_name = self.table_video.item(row, 0).text() + '.mp4'
                video_path = self.chanel_name + f'/{video_name}'
                if os.path.exists(video_path):
                    for col in range(columns - 2):
                        self.table_video.item(row, col).setBackground(QColor('#037400'))

        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

    def auto_check(self, table, state):

        columns = table.columnCount()
        rows = table.rowCount()
        if table == self.table_user:
            for row in range(rows):
                checkbox = table.cellWidget(row, columns - 1)
                if checkbox.checkState() != self.chack_box_states[checkbox]:
                    self.chack_box_states[checkbox] = checkbox.checkState()
                    columns_table_playlist = self.table_playlist.columnCount()
                    rows_table_playlist = self.table_playlist.rowCount()

                    if checkbox.checkState():
                        for col in range(columns - 1):
                            self.table_user.item(row, col).setBackground(QColor('#ffe5e5'))
                    else:
                        for col in range(columns - 1):
                            self.table_user.item(row, col).setBackground(QColor('#ffffff'))

                    for row_table_playlist in range(rows_table_playlist):
                        checkbox_table_playlist = self.table_playlist.cellWidget(row_table_playlist, columns_table_playlist - 1)
                        checkbox_table_playlist.setCheckState(checkbox.checkState())

        if table == self.table_playlist:
            for row in range(rows):
                checkbox = table.cellWidget(row, columns - 1)
                if checkbox.checkState() != self.chack_box_states[checkbox]:
                    self.chack_box_states[checkbox] = checkbox.checkState()
                    play_list_name = table.item(row, columns - 2).text()
                    columns_table_video = self.table_video.columnCount()
                    rows_table_video = self.table_video.rowCount()

                    if checkbox.checkState():
                        for col in range(columns - 1):
                            self.table_playlist.item(row, col).setBackground(QColor('#ffe5e5'))
                    else:
                        for col in range(columns - 1):
                            self.table_playlist.item(row, col).setBackground(QColor('#ffffff'))

                    for row_table_video in range(rows_table_video):
                        play_list_name_in_table_video = self.table_video.item(row_table_video, columns_table_video - 4).text()
                        if play_list_name == play_list_name_in_table_video:
                            checkbox_table_video = self.table_video.cellWidget(row_table_video, columns_table_video - 1)
                            checkbox_table_video.setCheckState(checkbox.checkState())

        if table == self.table_video:
            for row in range(rows):
                checkbox = table.cellWidget(row, columns - 1)
                if checkbox.checkState() != self.chack_box_states[checkbox]:
                    self.chack_box_states[checkbox] = checkbox.checkState()

                    if checkbox.checkState():
                        video_name = self.table_video.item(row, 0).text() + '.mp4'
                        video_path = self.chanel_name + f'/{video_name}'
                        if os.path.exists(video_path):
                            for col in range(columns - 2):
                                self.table_video.item(row, col).setBackground(QColor('#037400'))
                        else:
                            for col in range(columns - 2):
                                self.table_video.item(row, col).setBackground(QColor('#ffe5e5'))
                    else:
                        video_name = self.table_video.item(row, 0).text() + '.mp4'
                        video_path = self.chanel_name + f'/{video_name}'
                        if os.path.exists(video_path):
                            for col in range(columns - 2):
                                self.table_video.item(row, col).setBackground(QColor('#037400'))
                        else:
                            for col in range(columns - 2):
                                self.table_video.item(row, col).setBackground(QColor('#ffffff'))

class main_window(QMainWindow):

    @error_decorator
    def __init__(self):
        super().__init__()

        self.video_uploads = False

        self.setWindowIcon(QIcon('youtube_icon.jpg'))
        self.setWindowTitle('YouTubeBot')

        self.setFixedSize(1800, 900)

        self.tab_window = my_tab_widget(self)
        self.tab_window.setFixedSize(1800, 800)

        self.first_window = first_dialog()
        self.first_window.show()
        self.first_window.button.clicked.connect(self.show_main_window)
        self.first_window.text.returnPressed.connect(self.show_main_window)

        self.pb_upload_xl = QPushButton('Скачать в Excel', self)
        self.pb_upload_xl.move(810, 825)
        self.pb_upload_xl.setFixedSize(200, 50)
        self.pb_upload_xl.clicked.connect(self.info_from_account_in_excel)

        self.pb_upload_video = QPushButton('Скачать видео', self)
        self.pb_upload_video.setCheckable(True)
        self.pb_upload_video.move(590, 825)
        self.pb_upload_video.setFixedSize(200, 50)
        self.pb_upload_video.clicked.connect(self.upload_videos)

        self.setStyleSheet('''
            QPushButton{
                font-size: 20px;
                background: #ffffff;
                color: #ce2c2a;
                border: 2px solid #ce2c2a;
                border-radius: 6px;
                padding: 10px 0px 10px 0px;
            }
            QPushButton:hover{
                background: #ce2c2a;
                color: #ffffff;
            }
            QPushButton:checked{
                background: #ce2c2a;
                color: #ffffff;
            }
            QMainWindow{
                background: #ffffff;
                width: 10px; /* when vertical */
                height: 10px; /* when horizontal */
            }
            QProgressBar {
                border: 2px solid grey;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #037400;
                width: 2px;
            }
            QTabWidget::tab-bar {
                left: 5px; /* move to the right by 5px */
            }
            QTabBar::tab {
                background: #ffffff;
                color: #ce2c2a;
                font-size: 14px;
                border: 1px solid #ce2c2a;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                width: 20ex;
                padding: 5px 20px 5px 20px;
            }
            QTabBar::tab:selected {
                background: #ce2c2a;
                color: #ffffff;
                margin-left: -4px;
                margin-right: -4px;
            }
            QTabBar::tab:hover {
                background: #ce2c2a;
                color: #ffffff;
            }
            QTabBar::tab:!selected {
                margin-top: 2px; /* make non-selected tabs look smaller */
            }
            QTableWidget{
                background-color: #ffffff;
                selection-background-color: #ce2c2a;
                selection-color: #ffffff;
            }
            QTableWidget::item{
                padding: 5px;
            }
            QScrollBar:horizontal {
                background: #ec9e9d;
                height: 15px;
                margin: 0px 20px 0 0px;
            }
            QScrollBar::handle:horizontal {
                background: #ce2c2a;
                height: 15px;
                margin: 0px 20px 0 0px;
            }
            QScrollBar:vertical {
                 background: #ec9e9d;
                 width: 15px;
                 margin: 22px 0 22px 0;
             }
             QScrollBar::handle:vertical {
                 background: #ce2c2a;
                 min-height: 20px;
             }
            ''')

    @error_decorator
    def show_main_window(self, state_of_button=None):

        username = self.first_window.text.text()
        if username == '':
            return None
        else:
            self.get_info(username)
            self.first_window.close()
            self.show()

    @error_decorator
    def upload_videos(self, state_of_button):

        if not self.pb_upload_video.isChecked():
            self.pb_upload_video.toggle()
            return None

        columns = self.tab_window.table_video.columnCount()
        rows = self.tab_window.table_video.rowCount()
        
        for row in range(rows):
            if self.tab_window.table_video.cellWidget(row, columns - 1).isChecked():
                self.video_uploads = True
        
        if self.pb_upload_video.isChecked() and not self.video_uploads:
            self.pb_upload_video.toggle()
            return None

        list_of_video_urls = []

        for row in range(rows):
            if self.tab_window.table_video.cellWidget(row, columns - 1).isChecked():
                list_of_video_urls.append(self.tab_window.table_video.item(row, columns - 3).text())

        self.workers = WorkThread(list_of_video_urls, self.chanel_name)
        # запускает новый поток
        self.workers.start()
        # когда поступает сигнал, то вызывает функцию self.update_progress(self, video_url, persent)
        self.workers.update.connect(self.update_progress)
        self.workers.finish.connect(self.togle_upload_button)

    @error_decorator
    def get_info(self, username):

        self.worker = Parser(username)
        # запускает новый поток
        self.worker.start()
        # когда поступает сигнал, то вызывает функцию self.update_progress(self, video_url, persent)
        self.worker.info.connect(self.info_from_account_to_tables)

    @error_decorator
    def info_from_account_to_tables(self, chanel_info, play_lists_info, videos_info):

        self.chanel_info = chanel_info
        self.chanel_name = self.chanel_info['Название канала']

        self.play_lists_info = play_lists_info
        for play_list_number in self.play_lists_info:
            del self.play_lists_info[play_list_number]['play_list_id']

        self.videos_info = videos_info
        for video_number in self.videos_info:
            del self.videos_info[video_number]['video id']

        self.tab_window.info_to_table(self.tab_window.table_user, self.chanel_info, self.chanel_name)
        self.tab_window.info_to_table(self.tab_window.table_playlist, self.play_lists_info, self.chanel_name)
        self.tab_window.info_to_table(self.tab_window.table_video, self.videos_info, self.chanel_name)

    @error_decorator
    def update_progress(self, video_url, persent):

        rows = self.tab_window.table_video.rowCount()
        column = self.tab_window.table_video.columnCount()
        for row in range(rows):
            row_video_url = self.tab_window.table_video.item(row, column - 3).text()
            if row_video_url == video_url:
                self.tab_window.table_video.cellWidget(row, column - 2).setValue(int(persent))
                if int(persent) == 100:
                    for col in range(column - 2):
                        self.tab_window.table_video.item(row, col).setBackground(QColor('#037400'))

    @error_decorator
    def togle_upload_button(self, bool_unswer):
        if bool_unswer:
            self.pb_upload_video.toggle()
            self.video_uploads = False

    @error_decorator
    def info_from_account_in_excel(self, button_state=None):

        w = Excel(self.chanel_info, self.play_lists_info, self.videos_info)
        w.start()

class Excel(QThread):

    def __init__(self, chanel_info, play_lists_info, videos_info):
        super().__init__()

        self.chanel_info = chanel_info
        self.chanel_name = self.chanel_info['Название канала']
        self.play_lists_info = play_lists_info
        self.videos_info = videos_info

    def run(self):
        '''
        Создает файл excel в рабочей директории и загружает туда всю информацию.
        '''

        wb = Workbook()
        sheet_1 = wb.create_sheet('Аккаунт')
        sheet_2 = wb.create_sheet('Плейлисты')
        sheet_3 = wb.create_sheet('Видео')

        light_style = NamedStyle(name='light_style')
        light_style.font = Font(size=16)
        light_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        light_style.fill = PatternFill(fgColor="99CCFF", fill_type="solid")
        wb.add_named_style(light_style)

        dark_style = NamedStyle(name='dark_style')
        dark_style.font = Font(size=16)
        dark_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        dark_style.fill = PatternFill(fgColor="33CCCC", fill_type="solid")
        wb.add_named_style(dark_style)

        header_style = NamedStyle(name='header_style')
        header_style.font = Font(size=20, bold=True)
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        header_style.fill = PatternFill(fgColor="1896a5", fill_type="solid")
        wb.add_named_style(header_style)

        sheet_1.cell(row=1, column=1, value='Название канала')
        sheet_1.cell(row=1, column=2, value='Дата создания')
        sheet_1.cell(row=1, column=3, value='Страна')
        sheet_1.cell(row=1, column=4, value='Количество подписчиков')
        sheet_1.cell(row=1, column=5, value='Количество видео')
        sheet_1.cell(row=1, column=6, value='Количество просмотров')
        sheet_1.cell(row=1, column=7, value='Описание')

        sheet_1.cell(row=2, column=1, value=self.chanel_info['Название канала'])
        sheet_1.cell(row=2, column=2, value=self.chanel_info['Дата создания'].isoformat(' '))
        sheet_1.cell(row=2, column=3, value=self.chanel_info['Страна'])
        sheet_1.cell(row=2, column=4, value=self.chanel_info['Количество подписчиков'])
        sheet_1.cell(row=2, column=5, value=self.chanel_info['Количество видео'])
        sheet_1.cell(row=2, column=6, value=self.chanel_info['Количество просмотров'])
        sheet_1.cell(row=2, column=7, value=self.chanel_info['Описание'])

        sheet_2.cell(row=1, column=1, value='Название плейлиста')
        sheet_2.cell(row=1, column=2, value='Количество видео в плейлисте')
        sheet_2.cell(row=1, column=3, value='Дата публикации')
        sheet_2.cell(row=1, column=4, value='Описание')

        row = 2
        for play_list_number in self.play_lists_info:

            sheet_2.cell(row=row, column=1, value=self.play_lists_info[play_list_number]['Название плейлиста'])
            sheet_2.cell(row=row, column=2, value=self.play_lists_info[play_list_number]['Количество видео в плейлисте'])
            sheet_2.cell(row=row, column=3, value=self.play_lists_info[play_list_number]['Дата публикации'].isoformat(' '))
            sheet_2.cell(row=row, column=4, value=self.play_lists_info[play_list_number]['Описание'])

            row += 1

        sheet_3.cell(row=1, column=1, value='Название ролика')
        sheet_3.cell(row=1, column=2, value='Дата публикации')
        sheet_3.cell(row=1, column=3, value='Коментарии')
        sheet_3.cell(row=1, column=4, value='Дизлайки')
        sheet_3.cell(row=1, column=5, value='Лайки')
        sheet_3.cell(row=1, column=6, value='Просмотры')
        sheet_3.cell(row=1, column=7, value='Продолжительность Видео')
        sheet_3.cell(row=1, column=8, value='Плейлист')
        sheet_3.cell(row=1, column=9, value='Ссылка')

        row = 2
        for video_number in self.videos_info:

            sheet_3.cell(row=row, column=1, value=self.videos_info[video_number]['Название ролика'])
            sheet_3.cell(row=row, column=2, value=self.videos_info[video_number]['Дата публикации'].isoformat(' '))
            sheet_3.cell(row=row, column=3, value=self.videos_info[video_number]['Коментарии'])
            sheet_3.cell(row=row, column=4, value=self.videos_info[video_number]['Дизлайки'])
            sheet_3.cell(row=row, column=5, value=self.videos_info[video_number]['Лайки'])
            sheet_3.cell(row=row, column=6, value=self.videos_info[video_number]['Просмотры'])
            sheet_3.cell(row=row, column=7, value=self.videos_info[video_number]['Продолжительность Видео'].isoformat(timespec='auto'))
            sheet_3.cell(row=row, column=8, value=self.videos_info[video_number]['Плейлист'])
            sheet_3.cell(row=row, column=9, value='Посмотреть видео').hyperlink = self.videos_info[video_number]['Ссылка']

            row += 1

        for sheet in [sheet_1, sheet_2, sheet_3]:

            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells) + 16
                sheet.column_dimensions[column_cells[0].column_letter].width = length

            count = 1
            for row in sheet.rows:

                column_count = 0

                if count % 2 == 0:
                    for cell in row:
                        column_count += 1
                        cell.style = light_style
                        if column_count == 9:
                            cell.font = Font(size=16, color='d02c2a')

                if count == 1:
                    for cell in row:
                        cell.style = header_style

                elif count % 2 != 0:
                    for cell in row:
                        column_count += 1
                        cell.style = dark_style
                        if column_count == 9:
                            cell.font = Font(size=16, color='d02c2a')

                count += 1

            sheet.freeze_panes = 'A2'

        std = wb['Sheet']
        wb.remove(std)

        sheet_2.auto_filter.ref = sheet_2.dimensions
        sheet_3.auto_filter.ref = sheet_3.dimensions

        wb.save(self.chanel_info['Название канала'] + '.xlsx')

class Parser(QThread):

    info = pyqtSignal(dict, dict, dict)

    def __init__(self, username):

        super().__init__()

        self.username = username

        api_key = 'Ваш апи ключ для youtube'

        self.youtube = build('youtube', 'v3', developerKey=api_key)

        self.play_lists = {}
        self.play_lists_info = {}
        self.videos_info = {}

        request = self.youtube.channels().list(
            part='id',
            forUsername=f'{self.username}'
        )

        response = request.execute()

        self.chanel_id = response['items'][0]['id']

    def get_account_info(self):

        self.chanel_info = {}

        request = self.youtube.channels().list(
            part='statistics, contentDetails, contentOwnerDetails, snippet',
            id=f'{self.chanel_id}'
        )

        response = request.execute()

        chanel_info = json.dumps(response, sort_keys=True, indent=4, ensure_ascii=False)

        chanel_info = json.loads(chanel_info)

        statistics = chanel_info['items'][-1]["statistics"]
        snippet = chanel_info['items'][-1]["snippet"]

        self.chanel_info['Название канала'] = snippet['title']
        self.chanel_info['Дата создания'] = datetime.datetime.fromisoformat(snippet['publishedAt'][:-1])
        self.chanel_info['Страна'] = snippet['country']
        self.chanel_info['Описание'] = snippet['description']
        self.chanel_info['Количество подписчиков'] = int(statistics['subscriberCount'])
        self.chanel_info['Количество видео'] = int(statistics['videoCount'])
        self.chanel_info['Количество просмотров'] = int(statistics['viewCount'])

    def get_play_lists_from_account(self):

        next_page_token = None
        play_list_count = 1
        while True:

            request = self.youtube.playlists().list(
                part="snippet, contentDetails",
                channelId=f'{self.chanel_id}',
                maxResults=25,
                pageToken=next_page_token
            )

            response = request.execute()

            chanel_info = json.dumps(response, sort_keys=True, indent=4, ensure_ascii=False)

            chanel_info = json.loads(chanel_info)

            for play_list_info in chanel_info['items']:
                self.play_lists[f'{play_list_count}'] = play_list_info
                play_list_count += 1

            try:
                next_page_token = chanel_info['nextPageToken']
            except:
                break

    def get_play_lists_info(self):

        for play_list_number in self.play_lists:

            info = self.play_lists[play_list_number]

            self.play_lists_info[play_list_number] = {}
            self.play_lists_info[play_list_number]['play_list_id'] = info['id']
            self.play_lists_info[play_list_number]['Количество видео в плейлисте'] = int(info['contentDetails']['itemCount'])
            self.play_lists_info[play_list_number]['Описание'] = info['snippet']['description']
            self.play_lists_info[play_list_number]['Дата публикации'] = datetime.datetime.fromisoformat(info['snippet']['publishedAt'][:-1])
            self.play_lists_info[play_list_number]['Название плейлиста'] = info['snippet']['title']

    def get_videos_id(self):

        self.videos_id = {}

        next_page_token = None
        for play_list_number in self.play_lists_info:

            while True:

                play_list_id = self.play_lists_info[play_list_number]['play_list_id']

                request = self.youtube.playlistItems().list(
                    part="snippet, contentDetails",
                    playlistId=play_list_id,
                    maxResults=25,
                    pageToken=next_page_token
                )

                response = request.execute()

                chanel_info = json.dumps(response, sort_keys=True, indent=4, ensure_ascii=False)

                chanel_info = json.loads(chanel_info)

                for info in chanel_info['items']:
                    self.videos_id[info['contentDetails']['videoId']] = self.play_lists_info[play_list_number]['Название плейлиста']

                try:
                    next_page_token = chanel_info['nextPageToken']
                except:
                    next_page_token = None
                    break

    def get_videos_info(self):

        video_count = 0

        videos_id_strings = []

        step = 25
        min = 0
        max = step

        videos_id_list = []
        for video_id in self.videos_id:
            videos_id_list.append(video_id)

        for i in self.videos_id:

            videos_id = ','.join(videos_id_list[min:max])
            videos_id_strings.append(videos_id)

            min += step
            max += step

            if videos_id == '':
                break

        for videos_id_string in videos_id_strings:
            request = self.youtube.videos().list(
                    part="snippet,contentDetails,statistics",
                    id=videos_id_string
                )

            response = request.execute()

            chanel_info = json.dumps(response, sort_keys=True, indent=4, ensure_ascii=False)

            chanel_info = json.loads(chanel_info)

            hours_pattern = re.compile(r'(\d+)H')
            minutes_pattern = re.compile(r'(\d+)M')
            seconds_pattern = re.compile(r'(\d+)S')

            for item in chanel_info['items']:

                video_count += 1
                self.videos_info[str(video_count)] = {}

                duration = item['contentDetails']['duration']

                hours = hours_pattern.search(duration)
                minutes = minutes_pattern.search(duration)
                seconds = seconds_pattern.search(duration)

                hours = hours.group(1) if hours else '00'
                if len(hours) == 1:
                    hours = '0' + hours
                minutes = minutes.group(1) if minutes else '00'
                if len(minutes) == 1:
                    minutes = '0' + minutes
                seconds = seconds.group(1) if seconds else '00'
                if len(seconds) == 1:
                    seconds = '0' + seconds


                self.videos_info[str(video_count)]['Название ролика'] = item['snippet']['title']
                self.videos_info[str(video_count)]['Дата публикации'] = datetime.datetime.fromisoformat(item['snippet']['publishedAt'][:-1])
                try:
                    self.videos_info[str(video_count)]['Коментарии'] = int(item['statistics']['commentCount'])
                    self.videos_info[str(video_count)]['Дизлайки'] = int(item['statistics']['dislikeCount'])
                    self.videos_info[str(video_count)]['Лайки'] = int(item['statistics']['likeCount'])
                except:
                    self.videos_info[str(video_count)]['Коментарии'] = ''
                    self.videos_info[str(video_count)]['Дизлайки'] = ''
                    self.videos_info[str(video_count)]['Лайки'] = ''
                self.videos_info[str(video_count)]['Просмотры'] = int(item['statistics']['viewCount'])
                self.videos_info[str(video_count)]['Продолжительность Видео'] = datetime.time.fromisoformat(hours + ':' + minutes + ':' + seconds)
                self.videos_info[str(video_count)]['video id'] = item['id']
                self.videos_info[str(video_count)]['Плейлист'] = self.videos_id[item['id']]
                self.videos_info[str(video_count)]['Ссылка'] = f"https://www.youtube.com/watch?v={item['id']}"

    def run(self):

        self.get_account_info()
        self.get_play_lists_from_account()
        self.get_play_lists_info()
        self.get_videos_id()
        self.get_videos_info()

        self.info.emit(self.chanel_info, self.play_lists_info, self.videos_info)

class WorkThread(QThread):
    '''
    Класс которые представляет собой новый поток.
    '''

    # создает экземпляр сигнала
    update = pyqtSignal(str, float)
    finish = pyqtSignal(bool)

    def __init__(self, list_of_video_urls, chanel_name):
        super().__init__()

        self.chanel_name = chanel_name
        self.streams_urls = {}

        self.list_of_video_urls = list_of_video_urls
        self.max_thread = 3
        self.lock = Semaphore(self.max_thread)

    def upload_video_by_url(self, url):

        with self.lock:
            current_path = os.getcwd()
            main_folder_name = self.chanel_name
            path = current_path + f'\{main_folder_name}'
            if not os.path.exists(path):
                os.makedirs(path)

            yt = YouTube(url, on_progress_callback=self.return_progress_persent)
            self.video = yt.streams.filter(file_extension='mp4').first()
            self.streams_urls[self.video] = url
            try:
                self.video.download(path)
            except:
                pprint('Возникла ошибка во время скачивания видео')

            if url == self.list_of_video_urls[-1]:
                self.finish.emit(True)

    def return_progress_persent(self, stream, file_handle, bytes_remaining):

        video_size = stream.filesize
        persent = round((1-bytes_remaining/video_size)*100, 2)

        if persent % 1.0 == 0.0:
            self.update.emit(self.streams_urls[stream], persent)

    def run(self):

        for url in self.list_of_video_urls:
            time.sleep(0.5)
            tread = Thread(target=self.upload_video_by_url, args=(url, ))
            tread.start()

app = QApplication(sys.argv)
window = main_window()
app.exec_()
